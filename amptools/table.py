# stdlib imports
import re
import os.path
import time
import string

# third party imports
import pandas as pd
import numpy as np
from lxml import etree
from openpyxl import load_workbook, utils


REQUIRED_COLUMNS = ['station', 'lat', 'lon', 'netid']
CHANNEL_GROUPS = [['[a-z]{2}e', '[a-z]{2}n', '[a-z]{2}z'],
                  ['h1', 'h2', 'z'],
                  ['unk']]
PGM_COLS = ['pga', 'pgv', 'psa03', 'psa10', 'psa30']
OPTIONAL = ['name', 'distance', 'reference', 'intensity', 'source','loc','insttype','elev']


def _move(cellstr, nrows, ncols):
    """Internal method for getting new cell coordinate by adding rows/columns to cell coordinate.

    'A1' moved by 1 row, 1 column => 'B2'

    Args:
        cellstr (str): Cell coordinate (A1,B2)
        nrows (int): Number of rows to move (usually down)
        ncols (int): Number of columns to move (usually right)
    Returns:
        str: New cell coordinate.
    """
    # WARNING! This will only work up to column Z!
    # colidx is a string, rowidx is a number
    col_str_idx, rowidx = utils.coordinate_from_string(cellstr)
    letters = string.ascii_uppercase
    try:
        colidx = letters.index(col_str_idx)
        newcolidx = colidx + ncols
        newrowidx = rowidx + nrows
        newcellstr = '%s%i' % (letters[newcolidx], newrowidx)
        return newcellstr
    except ValueError:
        raise ValueError('Could not add %i columns to column %s.' %
                         (ncols, col_str_idx))


def read_excel(excelfile):
    """Read strong motion Excel spreadsheet, return a DataFrame.

    Args:
        excelfile (str): Path to valid Excel file.
    Returns:
        DataFrame: Multi-indexed dataframe as described below.
        str or None: Reference string or None.

     - "station" String containing UNIQUE identifying station information.
     - "lat" Latitude where peak ground motion observations were made.
     - "lon" Longitude where peak ground motion observations were made.
     - "netid" (usually) two letter code indicating the source network.

    Optional columns include:
     - "name" String describing area where peak ground motions were observed.
     - "source" String describing (usu. long form) source of peak ground motion data.
     - "distance" Distance from epicenter to station location, in units of km.
     - "loc" Two character location code.
     - "insttype" Instrument type, str.
     - "elev" Station elevation, in meters.

    And then at least one of the following columns:
     - "intensity" MMI value (1-10).

        AND/OR
      a grouped set of per-channel peak ground motion columns, like this:

      -------------------------------------------------------------------------------
      |         H1              |           H2            |             Z           |
      -------------------------------------------------------------------------------
      |pga|pgv|psa03|psa10|psa30|pga|pgv|psa03|psa10|psa30|pga|pgv|psa03|psa10|psa30|
      -------------------------------------------------------------------------------

      The peak ground motion columns can be any of the following:
      - "pga" Peak ground acceleration in units of %g.
      - "pgv" Peak ground velocity in units of cm/sec.
      - "psa03" Peak spectral acceleration at 0.3 seconds, in units of %g.
      - "psa10" Peak spectral acceleration at 1.0 seconds, in units of %g.
      - "psa30" Peak spectral acceleration at 3.0 seconds, in units of %g.

    Valid "channel" columns are {H1,H2,Z} or {XXN,XXE,XXZ}, where 'XX' is any
    two-letter combination, usually adhering to the following standard:
    http://www.fdsn.org/seed_manual/SEEDManual_V2.4_Appendix-A.pdf

    If the input data set provides no channel information, then the channel can be simply
    "UNK".

    """

    # figure out if data frame is multi-index or not
    wb = load_workbook(excelfile)
    ws = wb.active

    # figure out where the top left of the data begins
    topleft = 'A1'
    first_cell = 'A2'
    second_cell = 'A3'

    # figure out if there is a little reference section in this...
    reference = None
    skip_rows = None
    header = [0, 1]
    if ws[topleft].value.lower() != 'reference':
        raise KeyError('Reference cells are required in A1 and B1!')
    refcell = _move(topleft, 0, 1)
    reference = ws[refcell].value
    first_cell = _move(topleft, 1, 0)
    second_cell = _move(first_cell, 1, 0)
    skip_rows = [0]
    header = [1, 2]

    is_multi = True
    # if the first column of the second row is not empty,
    # then we do not have a multi-index.
    if ws[second_cell].value is not None:
        is_multi = False

    # read in dataframe, assuming that ground motions are grouped by channel
    if is_multi:
        try:
            df = pd.read_excel(excelfile, header=header)
        except pd.errors.ParserError:
            raise IndexError('Input file has invalid empty first data row.')

        headers = df.columns.get_level_values(0).str.lower()
        subheaders = df.columns.get_level_values(1).str.lower()
        df.columns = pd.MultiIndex.from_arrays([headers, subheaders])
        top_headers = df.columns.levels[0]
    else:
        df = pd.read_excel(excelfile, skiprows=skip_rows)
        top_headers = df.columns

    # make sure basic columns are present
    if 'station' not in top_headers:
        df['station'] = df.index
        top_headers = df.columns.levels[0]
    if not set(REQUIRED_COLUMNS).issubset(set(top_headers)):
        fmt = 'Input Excel file must specify the following columns: %s.'
        tpl = (str(REQUIRED_COLUMNS))
        raise KeyError(fmt % tpl)

    # check if channel headers are valid
    channels = (set(top_headers) - set(REQUIRED_COLUMNS)) - set(OPTIONAL)
    valid = False
    if len(channels):
        for channel_group in CHANNEL_GROUPS:
            num_channels = 0
            for channel_pat in channel_group:
                cp = re.compile(channel_pat)
                if len(list(filter(cp.match, channels))):
                    num_channels += 1
            if num_channels == 1 and len(channels) == 1:
                valid = True
                break
            elif num_channels > 1:
                h1_pat = re.compile(channel_group[0])
                h2_pat = re.compile(channel_group[1])
                has_h1 = len(list(filter(h1_pat.match, channels))) > 0
                has_h2 = len(list(filter(h2_pat.match, channels))) > 0
                if has_h1 or has_h2:
                    valid = True
                    break
    else:
        valid = True
    if not valid:
        raise KeyError('%s is not a valid channel grouping' %
                       str(sorted(list(channels))))

    # make sure the empty cells are all nans or floats
    found = False
    if 'intensity' in top_headers:
        found = True
    empty_cell = re.compile('\s+')
    for channel in channels:
        channel_df = df[channel].copy()
        for column in PGM_COLS:
            if column in channel_df:
                found = True
                channel_df[column] = channel_df[column].replace(
                    empty_cell, np.nan)
                channel_df[column] = channel_df[column].astype(float)
        df[channel] = channel_df

    if not found:
        fmt = 'File must contain at least one of the following data columns: %s'
        tpl = (str(PGM_COLS + ['intensity']))
        raise KeyError(fmt % tpl)

    return (df, reference)


def dataframe_to_xml(df, xmlfile, reference=None):
    """Write a dataframe to ShakeMap XML format.

    This method accepts either a dataframe from read_excel, or
    one with this structure:
     - station: Station code (REQUIRED)
     - channel: Channel (HHE,HHN, etc.) (REQUIRED)
     - imt: Intensity measure type (pga,pgv, etc.) (REQUIRED)
     - value: IMT value. (REQUIRED)
     - lat: Station latitude. (REQUIRED)
     - lon: Station longitude. (REQUIRED)
     - netid: Station contributing network. (REQUIRED)
     - flag: String quality flag, meaningful to contributing networks, 
             but ShakeMap ignores any station with a non-zero value. (REQUIRED)
     - elev: Elevation of station (m). (OPTIONAL)
     - name: String describing station. (OPTIONAL)
     - distance: Distance (km) from station to origin. (OPTIONAL)
     - loc: Description of location (i.e., "5 km south of Wellington") (OPTIONAL)
     - insttype: Instrument type (FBA, etc.) (OPTIONAL)

    Args:
        df (DataFrame): Pandas dataframe, as described in read_excel.
        xmlfile (str): Path to file where XML file should be written.
    """
    if hasattr(df.columns, 'levels'):
        top_headers = df.columns.levels[0]
        channels = (set(top_headers) - set(REQUIRED_COLUMNS)) - set(OPTIONAL)
    else:
        channels = []
    root = etree.Element('shakemap-data', code_version="3.5", map_version="3")

    create_time = int(time.time())
    stationlist = etree.SubElement(
        root, 'stationlist', created='%i' % create_time)
    if reference is not None:
        stationlist.attrib['reference'] = reference

    processed_stations = []
        
    for _, row in df.iterrows():
        tmprow = row.copy()
        if isinstance(tmprow.index, pd.core.indexes.multi.MultiIndex):
            tmprow.index = tmprow.index.droplevel(1)

        # assign required columns
        stationcode = str(tmprow['station']).strip()
        
        netid = tmprow['netid'].strip()
        if not stationcode.startswith(netid):
            stationcode = '%s.%s' % (netid, stationcode)

        # if this is a dataframe created by shakemap,
        # there will be multiple rows per station.
        # below we process all those rows at once,
        # so we need this bookkeeping to know that
        # we've already dealt with this station
        if stationcode in processed_stations:
            continue

        station = etree.SubElement(stationlist, 'station')

        station.attrib['code'] = stationcode
        station.attrib['lat'] = '%.4f' % tmprow['lat']
        station.attrib['lon'] = '%.4f' % tmprow['lon']

        # assign optional columns
        if 'name' in tmprow:
            station.attrib['name'] = tmprow['name'].strip()
        if 'netid' in tmprow:
            station.attrib['netid'] = tmprow['netid'].strip()
        if 'distance' in tmprow:
            station.attrib['dist'] = '%.1f' % tmprow['distance']
        if 'intensity' in tmprow:
            station.attrib['intensity'] = '%.1f' % tmprow['intensity']
        if 'source' in tmprow:
            station.attrib['source'] = tmprow['source'].strip()
        if 'loc' in tmprow:
            station.attrib['loc'] = tmprow['loc'].strip()
        if 'insttype' in tmprow:
            station.attrib['insttype'] = tmprow['insttype'].strip()
        if 'elev' in tmprow:
            station.attrib['elev'] = '%.1f' % tmprow['elev']
                
        if 'imt' not in tmprow.index:
            # sort channels by N,E,Z or H1,H2,Z
            channels = sorted(list(channels))

            for channel in channels:
                component = etree.SubElement(station, 'comp')
                component.attrib['name'] = channel.upper()

                # create sub elements out of any of the PGMs
                for pgm in ['pga', 'pgv', 'psa03', 'psa10', 'psa30']:
                    if pgm not in row[channel] or np.isnan(row[channel][pgm]):
                        continue
                    if pgm in row[channel]:
                        pgm_el = etree.SubElement(component, pgm)
                        pgm_el.attrib['flag'] = '0'
                        pgm_el.attrib['value'] = '%.4f' % row[channel][pgm]
            processed_stations.append(stationcode)
        else:
            # this file was created by a process that has imt/value columns
            # search the dataframe for all rows with this same station code
            scode = tmprow['station']
            station_rows = df[df['station'] == scode]

            # now we need to find all of the channels
            channels = station_rows['channel'].unique()
            for channel in channels:
                channel_rows = station_rows[station_rows['channel'] == channel]
                component = etree.SubElement(station, 'comp')
                component.attrib['name'] = channel.upper()
                for _,channel_row in channel_rows.iterrows():
                    pgm = channel_row['imt']
                    value = channel_row['value']
                    
                    pgm_el = etree.SubElement(component, pgm)
                    pgm_el.attrib['value'] = '%.4f' % value
                    pgm_el.attrib['flag'] = str(channel_row['flag'])
                    
                
            processed_stations.append(stationcode)

    tree = etree.ElementTree(root)
    tree.write(xmlfile, pretty_print=True)
